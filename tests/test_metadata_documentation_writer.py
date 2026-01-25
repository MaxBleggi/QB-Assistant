"""
Tests for MetadataDocumentationWriter - Report Info and Methodology sheets.

Verifies correct generation of documentation sheets with metadata, methodology
explanations, assumptions, excluded periods, and footnotes.
"""
import pytest
from datetime import datetime
from unittest.mock import Mock, MagicMock
from openpyxl import Workbook

from src.exporters.metadata_documentation_writer import MetadataDocumentationWriter
from src.models import ForecastScenarioModel, AnomalyAnnotationModel
from src.models.multi_scenario_forecast_result import MultiScenarioForecastResult


class TestMetadataDocumentationWriter:
    """Test suite for MetadataDocumentationWriter."""

    @pytest.fixture
    def mock_multi_scenario_result(self):
        """Create mock MultiScenarioForecastResult for testing."""
        result = Mock(spec=MultiScenarioForecastResult)
        result.created_at = '2026-01-15T14:30:00'
        result.forecast_horizon = 6
        result.client_id = 'ABC Corp'
        result.list_scenarios.return_value = ['Base', 'Conservative']
        return result

    @pytest.fixture
    def mock_forecast_scenario_model_list(self):
        """Create list of mock ForecastScenarioModel instances."""
        scenario1 = Mock(spec=ForecastScenarioModel)
        scenario1.scenario_name = 'Base'
        scenario1.description = 'Expected growth scenario'
        scenario1.parameters = {
            'growth_rate_revenue': 0.05,
            'collection_period_days': 30,
            'major_events': 'Product launch in Q2'
        }

        scenario2 = Mock(spec=ForecastScenarioModel)
        scenario2.scenario_name = 'Conservative'
        scenario2.description = 'Lower growth scenario'
        scenario2.parameters = {
            'growth_rate_revenue': 0.02,
            'collection_period_days': 45
        }

        return [scenario1, scenario2]

    @pytest.fixture
    def mock_anomaly_annotation_model(self):
        """Create mock AnomalyAnnotationModel with sample annotations."""
        model = Mock(spec=AnomalyAnnotationModel)
        model.get_annotations.return_value = [
            {
                'start_date': '2025-01-01',
                'end_date': '2025-01-31',
                'metric_name': 'revenue',
                'reason': 'Holiday spike',
                'exclude_from': ['trend_analysis']
            },
            {
                'start_date': '2025-06-01',
                'end_date': '2025-06-30',
                'metric_name': 'expenses',
                'reason': 'One-time marketing campaign',
                'exclude_from': 'baseline'
            }
        ]
        return model

    @pytest.fixture
    def mock_anomaly_annotation_model_empty(self):
        """Create mock AnomalyAnnotationModel with no annotations."""
        model = Mock(spec=AnomalyAnnotationModel)
        model.get_annotations.return_value = []
        return model

    @pytest.fixture
    def writer(self):
        """Create MetadataDocumentationWriter instance."""
        return MetadataDocumentationWriter()

    def test_metadata_documentation_writer_creates_report_info_sheet(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify Report Info sheet created with correct metadata fields."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        # Verify Report Info sheet exists
        assert 'Report Info' in writer.workbook.sheetnames

        ws = writer.workbook['Report Info']

        # Check for Report Metadata section
        assert ws['A1'].value == 'Report Metadata'
        # Should be bold - check font attribute
        assert ws['A1'].font.bold is True

    def test_metadata_documentation_writer_creates_methodology_sheet(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify Methodology sheet created with all required sections."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        # Verify Methodology sheet exists
        assert 'Methodology' in writer.workbook.sheetnames

        ws = writer.workbook['Methodology']

        # Scan sheet for required section headers
        found_sections = set()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in ['Methodology Summary', 'Confidence Intervals', 'Assumptions', 'Excluded Periods', 'Footnotes']:
                    found_sections.add(cell.value)

        # Verify all required sections present
        assert 'Methodology Summary' in found_sections
        assert 'Confidence Intervals' in found_sections
        assert 'Assumptions' in found_sections
        assert 'Excluded Periods' in found_sections
        assert 'Footnotes' in found_sections

    def test_report_info_displays_generation_date(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify generation date extracted from MultiScenarioForecastResult.created_at and displayed."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        ws = writer.workbook['Report Info']

        # Find Generated On row
        found_date = False
        for row in ws.iter_rows():
            if row[0].value == 'Generated On:':
                # Check that date is displayed in column B
                date_value = row[1].value
                assert date_value is not None
                # Should contain date formatted from 2026-01-15
                assert '2026-01-15' in str(date_value)
                found_date = True
                break

        assert found_date, "Generated On field not found in Report Info sheet"

    def test_report_info_displays_forecast_horizon(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify forecast horizon (6 or 12 months) displayed correctly."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        ws = writer.workbook['Report Info']

        # Find Forecast Horizon row
        found_horizon = False
        for row in ws.iter_rows():
            if row[0].value == 'Forecast Horizon:':
                horizon_value = row[1].value
                assert horizon_value == '6 months'
                found_horizon = True
                break

        assert found_horizon, "Forecast Horizon field not found in Report Info sheet"

    def test_report_info_lists_scenarios(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify all scenario names from list_scenarios() displayed."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        ws = writer.workbook['Report Info']

        # Find Scenarios row
        found_scenarios = False
        for row in ws.iter_rows():
            if row[0].value == 'Scenarios:':
                scenarios_value = row[1].value
                assert scenarios_value is not None
                # Should contain both scenario names
                assert 'Base' in scenarios_value
                assert 'Conservative' in scenarios_value
                found_scenarios = True
                break

        assert found_scenarios, "Scenarios field not found in Report Info sheet"

    def test_methodology_includes_confidence_interval_guide(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify confidence interval interpretation guide with example text present."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        ws = writer.workbook['Methodology']

        # Scan for confidence interval interpretation text
        found_interpretation = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and '80% confidence' in str(cell.value):
                    # Check for full interpretation text
                    if 'fall between Lower and Upper bounds' in str(cell.value):
                        found_interpretation = True
                        break

        assert found_interpretation, "Confidence interval interpretation guide not found in Methodology sheet"

    def test_assumptions_section_displays_scenario_parameters(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify each scenario's parameters extracted and displayed."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        ws = writer.workbook['Methodology']

        # Scan for scenario names and parameters
        found_base_scenario = False
        found_conservative_scenario = False
        found_growth_rate = False

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    value_str = str(cell.value)
                    if 'Scenario: Base' in value_str:
                        found_base_scenario = True
                    if 'Scenario: Conservative' in value_str:
                        found_conservative_scenario = True
                    # Check for formatted growth rate (5% or 2%)
                    if 'Growth Rate Revenue' in value_str or '5.00%' in value_str or '2.00%' in value_str:
                        found_growth_rate = True

        assert found_base_scenario, "Base scenario not found in Assumptions section"
        assert found_conservative_scenario, "Conservative scenario not found in Assumptions section"
        assert found_growth_rate, "Growth rate parameters not found in Assumptions section"

    def test_excluded_periods_section_displays_annotations(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify anomaly annotations extracted and displayed with dates, reasons."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        ws = writer.workbook['Methodology']

        # Scan for excluded periods table headers and data
        found_table_header = False
        found_holiday_spike = False
        found_revenue_metric = False

        for row in ws.iter_rows():
            row_values = [str(cell.value) if cell.value else '' for cell in row]
            row_text = ' '.join(row_values)

            # Check for table headers
            if 'Start Date' in row_text and 'End Date' in row_text and 'Metric' in row_text:
                found_table_header = True

            # Check for annotation data
            if 'Holiday spike' in row_text:
                found_holiday_spike = True
            if 'revenue' in row_text.lower():
                found_revenue_metric = True

        assert found_table_header, "Excluded Periods table header not found"
        assert found_holiday_spike, "Holiday spike annotation reason not found"
        assert found_revenue_metric, "Revenue metric not found in excluded periods"

    def test_excluded_periods_handles_empty_annotations(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model_empty
    ):
        """Verify 'No periods excluded' message when get_annotations() returns empty list."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model_empty
        )

        ws = writer.workbook['Methodology']

        # Scan for "No periods excluded" message
        found_empty_message = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and 'No periods excluded' in str(cell.value):
                    found_empty_message = True
                    break

        assert found_empty_message, "'No periods excluded' message not found when annotations list is empty"

    def test_footnotes_section_includes_key_terms(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify footnotes section contains common term explanations (MoM, P&L, etc.)."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        ws = writer.workbook['Methodology']

        # Scan for key terms in footnotes
        found_terms = set()
        key_terms = ['MoM', 'P&L', 'Confidence Interval', 'Growth Rate', 'Scenario']

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    value_str = str(cell.value)
                    for term in key_terms:
                        if term in value_str:
                            found_terms.add(term)

        # Verify at least 5 common terms present
        assert len(found_terms) >= 5, f"Expected at least 5 key terms in footnotes, found {len(found_terms)}: {found_terms}"

    def test_format_bold_applied_to_section_headers(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify section headers use format_bold() for emphasis."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        # Check Report Info sheet headers
        ws_info = writer.workbook['Report Info']
        # Report Metadata header should be bold
        assert ws_info['A1'].font.bold is True

        # Check Methodology sheet headers
        ws_method = writer.workbook['Methodology']

        # Find section headers and verify bold formatting
        section_headers = ['Methodology Summary', 'Confidence Intervals', 'Assumptions', 'Excluded Periods', 'Footnotes']
        bold_headers_found = 0

        for row in ws_method.iter_rows():
            for cell in row:
                if cell.value in section_headers:
                    if cell.font and cell.font.bold:
                        bold_headers_found += 1

        # Should find at least 4 bold section headers
        assert bold_headers_found >= 4, f"Expected at least 4 bold section headers, found {bold_headers_found}"

    def test_auto_adjust_column_widths_called(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify auto_adjust_column_widths() called for both sheets."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        # Verify column widths were adjusted (should be > default 10)
        ws_info = writer.workbook['Report Info']
        ws_method = writer.workbook['Methodology']

        # Check that at least one column in each sheet has non-default width
        info_widths = [ws_info.column_dimensions[col].width for col in ['A', 'B', 'C'] if ws_info.column_dimensions[col].width]
        method_widths = [ws_method.column_dimensions[col].width for col in ['A', 'B', 'C'] if ws_method.column_dimensions[col].width]

        assert len(info_widths) > 0, "No column widths set in Report Info sheet"
        assert len(method_widths) > 0, "No column widths set in Methodology sheet"

    def test_write_orchestrates_both_sheets(
        self,
        writer,
        mock_multi_scenario_result,
        mock_forecast_scenario_model_list,
        mock_anomaly_annotation_model
    ):
        """Verify write() method calls write_report_info_sheet() and write_methodology_sheet()."""
        writer.write(
            mock_multi_scenario_result,
            mock_forecast_scenario_model_list,
            mock_anomaly_annotation_model
        )

        # Verify both sheets exist in workbook
        assert 'Report Info' in writer.workbook.sheetnames
        assert 'Methodology' in writer.workbook.sheetnames

        # Verify sheets are in expected order (Report Info first)
        sheet_names = writer.workbook.sheetnames
        report_info_index = sheet_names.index('Report Info')
        methodology_index = sheet_names.index('Methodology')
        assert report_info_index < methodology_index, "Report Info should be created before Methodology"

    def test_metadata_documentation_writer_registration(self):
        """Verify MetadataDocumentationWriter importable from src.exporters."""
        # This test verifies the registration in __init__.py
        from src.exporters import MetadataDocumentationWriter

        # Should be able to instantiate
        writer = MetadataDocumentationWriter()
        assert writer is not None
        assert hasattr(writer, 'write')
        assert hasattr(writer, 'workbook')
