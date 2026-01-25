"""
Unit tests for BaseExcelWriter formatting utilities and workbook management.

Tests formatting methods (currency, percentage, borders, headers), hierarchy
traversal, workbook save/load, and column width adjustment.
"""
import pytest
import tempfile
import os
from openpyxl import load_workbook

from src.exporters import BaseExcelWriter


class TestBaseExcelWriter:
    """Test suite for BaseExcelWriter class."""

    @pytest.fixture
    def sample_workbook(self):
        """Create BaseExcelWriter instance with sample workbook."""
        writer = BaseExcelWriter()
        # Create a test sheet
        ws = writer.workbook.create_sheet('Test Sheet')
        return writer, ws

    @pytest.fixture
    def sample_hierarchy_simple(self):
        """Create simple hierarchy dict for traversal testing."""
        return {
            'name': 'Root',
            'values': {'Jan': 1000, 'Feb': 2000},
            'children': [
                {
                    'name': 'Child1',
                    'values': {'Jan': 500, 'Feb': 600}
                },
                {
                    'name': 'Child2',
                    'values': {'Jan': 300, 'Feb': 400}
                }
            ]
        }

    @pytest.fixture
    def sample_hierarchy_nested(self):
        """Create 3-level hierarchy dict for traversal testing."""
        return {
            'name': 'Root',
            'values': {'Jan': 5000},
            'children': [
                {
                    'name': 'Child1',
                    'values': {'Jan': 3000},
                    'children': [
                        {
                            'name': 'Grandchild',
                            'values': {'Jan': 1000}
                        }
                    ]
                },
                {
                    'name': 'Child2',
                    'values': {'Jan': 2000}
                }
            ]
        }

    @pytest.fixture
    def sample_hierarchy_forecast(self):
        """Create hierarchy with three parallel value dicts for forecast testing."""
        return {
            'name': 'Revenue',
            'projected': {'Jan': 10000, 'Feb': 12000},
            'lower_bound': {'Jan': 9000, 'Feb': 11000},
            'upper_bound': {'Jan': 11000, 'Feb': 13000},
            'children': [
                {
                    'name': 'Product Sales',
                    'projected': {'Jan': 7000, 'Feb': 8000},
                    'lower_bound': {'Jan': 6500, 'Feb': 7500},
                    'upper_bound': {'Jan': 7500, 'Feb': 8500}
                }
            ]
        }

    def test_workbook_creation(self):
        """Test BaseExcelWriter creates valid openpyxl Workbook."""
        writer = BaseExcelWriter()

        assert writer.workbook is not None
        assert hasattr(writer.workbook, 'create_sheet')
        assert hasattr(writer.workbook, 'save')

    def test_header_style_application(self, sample_workbook):
        """Test apply_header_style() sets bold font, fill, and border."""
        writer, ws = sample_workbook

        # Write a value and apply header style
        ws['A1'] = 'Header'
        writer.apply_header_style(ws, 'A1')

        cell = ws['A1']
        assert cell.font.bold is True
        assert cell.fill.start_color.rgb == '00D9E1F2'
        assert cell.border.top.style == 'thin'
        assert cell.border.bottom.style == 'thin'
        assert cell.border.left.style == 'thin'
        assert cell.border.right.style == 'thin'

    def test_header_style_range(self, sample_workbook):
        """Test apply_header_style() works on cell ranges."""
        writer, ws = sample_workbook

        # Write values and apply header style to range
        ws['A1'] = 'Header1'
        ws['B1'] = 'Header2'
        ws['C1'] = 'Header3'
        writer.apply_header_style(ws, 'A1:C1')

        # Check all cells in range have header style
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}1']
            assert cell.font.bold is True
            assert cell.border.top.style == 'thin'

    def test_currency_formatting(self, sample_workbook):
        """Test format_currency() applies $#,##0.00 format."""
        writer, ws = sample_workbook

        # Test positive value
        ws['B2'] = 1234.56
        writer.format_currency(ws, 'B2')
        assert ws['B2'].number_format == '$#,##0.00'

        # Test negative value
        ws['B3'] = -5678.90
        writer.format_currency(ws, 'B3')
        assert ws['B3'].number_format == '$#,##0.00'

        # Test zero
        ws['B4'] = 0
        writer.format_currency(ws, 'B4')
        assert ws['B4'].number_format == '$#,##0.00'

    def test_currency_formatting_none(self, sample_workbook):
        """Test format_currency() handles None values gracefully."""
        writer, ws = sample_workbook

        ws['B2'] = None
        writer.format_currency(ws, 'B2')
        # Should not raise error, and number_format should remain default
        assert ws['B2'].value is None

    def test_percentage_formatting(self, sample_workbook):
        """Test format_percentage() applies 0.00% format."""
        writer, ws = sample_workbook

        ws['E2'] = 0.1523
        writer.format_percentage(ws, 'E2')
        assert ws['E2'].number_format == '0.00%'

        ws['E3'] = 0.05
        writer.format_percentage(ws, 'E3')
        assert ws['E3'].number_format == '0.00%'

    def test_border_application(self, sample_workbook):
        """Test apply_borders() creates borders around ranges."""
        writer, ws = sample_workbook

        # Create a range of cells
        for row in range(1, 4):
            for col in range(1, 4):
                ws.cell(row=row, column=col, value=f'Cell{row}{col}')

        # Apply borders to range
        writer.apply_borders(ws, 'A1:C3')

        # Check that borders are applied to all cells
        for row in range(1, 4):
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                assert cell.border.top.style == 'thin'
                assert cell.border.bottom.style == 'thin'
                assert cell.border.left.style == 'thin'
                assert cell.border.right.style == 'thin'

    def test_auto_column_width(self, sample_workbook):
        """Test auto_adjust_column_widths() sets appropriate widths."""
        writer, ws = sample_workbook

        # Create columns with different content lengths
        ws['A1'] = 'Short'
        ws['B1'] = 'Medium Length'
        ws['C1'] = 'Very Long Text Content That Should Make Column Wide'

        # Apply auto-width adjustment
        writer.auto_adjust_column_widths(ws)

        # Check that column C is wider than A
        width_a = ws.column_dimensions['A'].width
        width_c = ws.column_dimensions['C'].width

        assert width_c > width_a
        # Column C should be at least as wide as the text length * 1.2
        assert width_c >= len('Very Long Text Content That Should Make Column Wide') * 1.2

    def test_save_workbook(self):
        """Test save() creates .xlsx file loadable by openpyxl."""
        writer = BaseExcelWriter()
        ws = writer.workbook.create_sheet('Test')
        ws['A1'] = 'Test Data'

        # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Save workbook
            writer.save(tmp_path)

            # Verify file exists
            assert os.path.exists(tmp_path)

            # Load workbook to verify it's valid
            loaded_wb = load_workbook(tmp_path)
            assert 'Test' in loaded_wb.sheetnames
            assert loaded_wb['Test']['A1'].value == 'Test Data'
        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    def test_hierarchy_traversal_simple(self, sample_hierarchy_simple):
        """Test traverse_hierarchy() yields correct tuples for simple tree."""
        writer = BaseExcelWriter()

        results = list(writer.traverse_hierarchy(sample_hierarchy_simple))

        # Should yield: Root (level 0), Child1 (level 1), Child2 (level 1)
        assert len(results) == 3

        # Check root
        assert results[0][0] == 0  # indent level
        assert results[0][1] == 'Root'  # name
        assert results[0][2] == {'Jan': 1000, 'Feb': 2000}  # values

        # Check Child1
        assert results[1][0] == 1
        assert results[1][1] == 'Child1'
        assert results[1][2] == {'Jan': 500, 'Feb': 600}

        # Check Child2
        assert results[2][0] == 1
        assert results[2][1] == 'Child2'
        assert results[2][2] == {'Jan': 300, 'Feb': 400}

    def test_hierarchy_traversal_nested(self, sample_hierarchy_nested):
        """Test traverse_hierarchy() handles nested children correctly."""
        writer = BaseExcelWriter()

        results = list(writer.traverse_hierarchy(sample_hierarchy_nested))

        # Should yield: Root (0), Child1 (1), Grandchild (2), Child2 (1)
        assert len(results) == 4

        # Check indent levels
        assert results[0][0] == 0  # Root
        assert results[1][0] == 1  # Child1
        assert results[2][0] == 2  # Grandchild
        assert results[3][0] == 1  # Child2

        # Check grandchild
        assert results[2][1] == 'Grandchild'
        assert results[2][2] == {'Jan': 1000}

    def test_hierarchy_traversal_forecast(self, sample_hierarchy_forecast):
        """Test traverse_hierarchy() yields three value dicts for forecasts."""
        writer = BaseExcelWriter()

        results = list(writer.traverse_hierarchy(sample_hierarchy_forecast, forecast=True))

        # Should yield 2 rows: Revenue and Product Sales
        assert len(results) == 2

        # Check Revenue (parent)
        assert results[0][0] == 0  # indent level
        assert results[0][1] == 'Revenue'  # name
        assert results[0][2] == {'Jan': 10000, 'Feb': 12000}  # projected
        assert results[0][3] == {'Jan': 9000, 'Feb': 11000}  # lower_bound
        assert results[0][4] == {'Jan': 11000, 'Feb': 13000}  # upper_bound

        # Check Product Sales (child)
        assert results[1][0] == 1  # indent level
        assert results[1][1] == 'Product Sales'
        assert results[1][2] == {'Jan': 7000, 'Feb': 8000}  # projected
        assert results[1][3] == {'Jan': 6500, 'Feb': 7500}  # lower_bound
        assert results[1][4] == {'Jan': 7500, 'Feb': 8500}  # upper_bound

    def test_hierarchy_traversal_no_children(self):
        """Test traverse_hierarchy() handles node with no 'children' key."""
        writer = BaseExcelWriter()

        # Hierarchy without children key
        hierarchy = {
            'name': 'Leaf',
            'values': {'Jan': 100}
        }

        results = list(writer.traverse_hierarchy(hierarchy))

        # Should yield just the leaf node
        assert len(results) == 1
        assert results[0][1] == 'Leaf'
        assert results[0][2] == {'Jan': 100}

    def test_hierarchy_traversal_empty_children(self):
        """Test traverse_hierarchy() handles node with empty children list."""
        writer = BaseExcelWriter()

        hierarchy = {
            'name': 'Parent',
            'values': {'Jan': 500},
            'children': []
        }

        results = list(writer.traverse_hierarchy(hierarchy))

        # Should yield just the parent
        assert len(results) == 1
        assert results[0][1] == 'Parent'
