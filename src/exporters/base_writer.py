"""
Base Excel writer providing shared workbook management and formatting utilities.

All sheet-specific writers inherit from BaseExcelWriter to access consistent
formatting methods for currency, percentages, borders, headers, and auto-sizing.
"""
from typing import Any, Dict, Optional, Tuple, Generator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet


class BaseExcelWriter:
    """
    Base class for Excel export functionality with reusable formatting utilities.

    Manages an openpyxl Workbook instance and provides methods for applying
    professional formatting including headers, currency, percentages, borders,
    and automatic column width adjustment.
    """

    def __init__(self):
        """Initialize with new openpyxl Workbook."""
        self.workbook = Workbook()
        # Remove default sheet created by openpyxl
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

    def apply_header_style(self, ws: Worksheet, cell_range: str) -> None:
        """
        Apply header styling to cell or range (bold font, fill color, border).

        Args:
            ws: Worksheet containing cells
            cell_range: Cell address (e.g., 'A1') or range (e.g., 'A1:E1')
        """
        # Parse range to handle both single cells and ranges
        if ':' in cell_range:
            # Range like 'A1:E1'
            cells = ws[cell_range]
            # Flatten if it's a tuple of tuples
            if isinstance(cells, tuple) and isinstance(cells[0], tuple):
                cells = [cell for row in cells for cell in row]
            elif isinstance(cells, tuple):
                cells = list(cells)
        else:
            # Single cell like 'A1'
            cells = [ws[cell_range]]

        # Apply styles to each cell
        for cell in cells:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def format_currency(self, ws: Worksheet, cell_range: str) -> None:
        """
        Apply currency number format ($#,##0.00) to cell or range.

        Args:
            ws: Worksheet containing cells
            cell_range: Cell address (e.g., 'B2') or range (e.g., 'B2:D10')
        """
        # Parse range
        if ':' in cell_range:
            cells = ws[cell_range]
            # Flatten if needed
            if isinstance(cells, tuple) and isinstance(cells[0], tuple):
                cells = [cell for row in cells for cell in row]
            elif isinstance(cells, tuple):
                cells = list(cells)
        else:
            cells = [ws[cell_range]]

        # Apply currency format
        for cell in cells:
            if cell.value is not None:
                cell.number_format = '$#,##0.00'

    def format_percentage(self, ws: Worksheet, cell_range: str) -> None:
        """
        Apply percentage number format (0.00%) to cell or range.

        Args:
            ws: Worksheet containing cells
            cell_range: Cell address (e.g., 'E2') or range (e.g., 'E2:E10')
        """
        # Parse range
        if ':' in cell_range:
            cells = ws[cell_range]
            # Flatten if needed
            if isinstance(cells, tuple) and isinstance(cells[0], tuple):
                cells = [cell for row in cells for cell in row]
            elif isinstance(cells, tuple):
                cells = list(cells)
        else:
            cells = [ws[cell_range]]

        # Apply percentage format
        for cell in cells:
            if cell.value is not None:
                cell.number_format = '0.00%'

    def apply_borders(self, ws: Worksheet, cell_range: str) -> None:
        """
        Apply thin borders to all cells in range.

        Args:
            ws: Worksheet containing cells
            cell_range: Cell range (e.g., 'A1:E10')
        """
        cells = ws[cell_range]

        # Handle different cell range structures
        if isinstance(cells, tuple):
            # Multi-row range
            if isinstance(cells[0], tuple):
                cells = [cell for row in cells for cell in row]
            # Single row range
            else:
                cells = list(cells)

        # Apply borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in cells:
            cell.border = thin_border

    def auto_adjust_column_widths(self, ws: Worksheet) -> None:
        """
        Automatically adjust column widths based on content length.

        Args:
            ws: Worksheet to adjust
        """
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                if column_letter is None:
                    column_letter = cell.column_letter

                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Set adjusted width (multiply by 1.2 for padding)
            if column_letter:
                adjusted_width = max(max_length * 1.2, 10)  # Minimum width of 10
                ws.column_dimensions[column_letter].width = adjusted_width

    def traverse_hierarchy(
        self,
        hierarchy: Dict[str, Any],
        indent_level: int = 0,
        forecast: bool = False
    ) -> Generator[Tuple, None, None]:
        """
        Recursively traverse hierarchy tree and yield (indent_level, name, values) tuples.

        For regular models (BudgetModel), yields (indent_level, name, values_dict).
        For forecast models, yields (indent_level, name, projected, lower_bound, upper_bound).

        Args:
            hierarchy: Hierarchy dict with optional 'name', 'values', 'children' keys
            indent_level: Current indentation level (starts at 0)
            forecast: If True, extract three parallel value dicts (projected, lower_bound, upper_bound)

        Yields:
            For regular models: (indent_level, name, values_dict)
            For forecast models: (indent_level, name, projected_dict, lower_bound_dict, upper_bound_dict)
        """
        if not isinstance(hierarchy, dict):
            return

        # Get node name
        name = hierarchy.get('name', '')

        # Yield current node
        if name:
            if forecast:
                # Extract three parallel value dicts for forecast models
                projected = hierarchy.get('projected', {})
                lower_bound = hierarchy.get('lower_bound', {})
                upper_bound = hierarchy.get('upper_bound', {})
                yield (indent_level, name, projected, lower_bound, upper_bound)
            else:
                # Extract single values dict for regular models
                values = hierarchy.get('values', {})
                yield (indent_level, name, values)

        # Recursively process children
        children = hierarchy.get('children', [])
        for child in children:
            yield from self.traverse_hierarchy(child, indent_level + 1, forecast)

    def apply_trend_indicator(self, ws: Worksheet, cell: str, growth_rate: Optional[float]) -> None:
        """
        Apply trend indicator symbol and color to a cell based on growth rate.

        Applies:
        - '▲' with green font for positive growth
        - '▼' with red font for negative growth
        - '—' with gray font for zero or None growth

        Args:
            ws: Worksheet containing the cell
            cell: Cell address (e.g., 'C5')
            growth_rate: Growth rate as decimal (e.g., 0.05 for 5% growth, -0.03 for -3% decline)
        """
        if growth_rate is None or growth_rate == 0.0:
            # Zero or missing growth - neutral indicator
            ws[cell].value = '—'
            ws[cell].font = Font(color='808080')  # Gray RGB(128, 128, 128)
        elif growth_rate > 0:
            # Positive growth - upward indicator
            ws[cell].value = '▲'
            ws[cell].font = Font(color='00B050')  # Green RGB(0, 176, 80)
        else:
            # Negative growth - downward indicator
            ws[cell].value = '▼'
            ws[cell].font = Font(color='C00000')  # Red RGB(192, 0, 0)

    def apply_conditional_highlight(self, ws: Worksheet, cell_range: str, fill_color: str) -> None:
        """
        Apply background color fill to cell or range.

        Args:
            ws: Worksheet containing cells
            cell_range: Cell address (e.g., 'B5') or range (e.g., 'B5:B10')
            fill_color: Hex color code (e.g., 'FFE699' for yellow RGB(255, 230, 153))
        """
        # Parse range
        if ':' in cell_range:
            cells = ws[cell_range]
            # Flatten if needed
            if isinstance(cells, tuple) and isinstance(cells[0], tuple):
                cells = [cell for row in cells for cell in row]
            elif isinstance(cells, tuple):
                cells = list(cells)
        else:
            cells = [ws[cell_range]]

        # Apply fill color to each cell
        for cell in cells:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    def format_bold(self, ws: Worksheet, cell_range: str) -> None:
        """
        Apply bold font to cell or range.

        Args:
            ws: Worksheet containing cells
            cell_range: Cell address (e.g., 'A1') or range (e.g., 'A1:A3')
        """
        # Parse range
        if ':' in cell_range:
            cells = ws[cell_range]
            # Flatten if needed
            if isinstance(cells, tuple) and isinstance(cells[0], tuple):
                cells = [cell for row in cells for cell in row]
            elif isinstance(cells, tuple):
                cells = list(cells)
        else:
            cells = [ws[cell_range]]

        # Apply bold font to each cell
        for cell in cells:
            cell.font = Font(bold=True)

    def save(self, file_path: str) -> None:
        """
        Save workbook to .xlsx file.

        Args:
            file_path: Path where .xlsx file should be saved
        """
        self.workbook.save(file_path)
