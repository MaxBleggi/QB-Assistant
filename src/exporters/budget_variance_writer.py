"""
Budget vs Actual variance Excel writer with conditional highlighting.

Generates Budget vs Actual sheet comparing budgeted amounts with actual results,
showing Budget, Actual, Variance ($), and Variance (%) columns with conditional
formatting for unfavorable variances (red >10%, yellow 5-10%).
"""
from .base_writer import BaseExcelWriter
from openpyxl.styles import Alignment


class BudgetVarianceReportWriter(BaseExcelWriter):
    """
    Excel writer for Budget vs Actual variance report with conditional highlighting.

    Creates a sheet with four columns per period: Budget, Actual, Variance ($), and
    Variance (%). Applies red highlighting for unfavorable variances >10% and yellow
    for 5-10% unfavorable variances. Includes subtotals from calculated_rows.
    """

    def write(self, variance_model) -> None:
        """
        Generate Budget vs Actual sheet from VarianceModel.

        Args:
            variance_model: VarianceModel instance with variance hierarchy and calculated_rows
        """
        # Create Budget vs Actual sheet
        ws = self.workbook.create_sheet('Budget vs Actual')

        # Get periods from variance model (assuming single period for variance)
        # Extract first available period from hierarchy
        periods = self._extract_periods(variance_model.hierarchy)
        if not periods:
            ws['A1'] = 'No data available'
            return

        # Use first period (Epic 3 design is single period)
        period = periods[0]

        # Write header row
        ws['A1'] = 'Account'
        ws['B1'] = 'Budget'
        ws['C1'] = 'Actual'
        ws['D1'] = 'Variance ($)'
        ws['E1'] = 'Variance (%)'

        # Apply header style
        self.apply_header_style(ws, 'A1:E1')

        # Track current row
        current_row = 2

        # Traverse variance hierarchy and write rows
        # Hierarchy is dict with section keys ('Income', 'Expenses'), iterate over sections
        for section_name, section_node in variance_model.hierarchy.items():
            for indent_level, name, values in self.traverse_hierarchy(section_node):
                # Extract variance data for the period
                period_data = values.get(period, {})

                budget_value = period_data.get('budget_value')
                actual_value = period_data.get('actual_value')
                dollar_variance = period_data.get('dollar_variance')
                pct_variance = period_data.get('pct_variance')
                is_favorable = period_data.get('is_favorable')
                is_flagged = period_data.get('is_flagged')

                # Write account name with indentation
                cell = ws.cell(row=current_row, column=1, value=name)
                cell.alignment = Alignment(indent=indent_level)

                # Write budget value
                if budget_value is not None:
                    ws.cell(row=current_row, column=2, value=budget_value)

                # Write actual value
                if actual_value is not None:
                    ws.cell(row=current_row, column=3, value=actual_value)

                # Write dollar variance
                if dollar_variance is not None:
                    ws.cell(row=current_row, column=4, value=dollar_variance)

                # Write percentage variance
                if pct_variance is not None:
                    ws.cell(row=current_row, column=5, value=pct_variance)

                    # Apply conditional highlighting for unfavorable variances
                    if is_flagged and not is_favorable:
                        # Determine color based on variance magnitude
                        abs_pct = abs(pct_variance)
                        if abs_pct > 0.10:  # >10% unfavorable
                            # Red highlighting for variance $ and variance %
                            self.apply_conditional_highlight(ws, f'D{current_row}', 'FFC7CE')
                            self.apply_conditional_highlight(ws, f'E{current_row}', 'FFC7CE')
                        elif abs_pct >= 0.05:  # 5-10% unfavorable
                            # Yellow highlighting for variance $ and variance %
                            self.apply_conditional_highlight(ws, f'D{current_row}', 'FFE699')
                            self.apply_conditional_highlight(ws, f'E{current_row}', 'FFE699')

                current_row += 1

        # Write calculated rows (subtotals for Revenue, Expenses, Net Income)
        if variance_model.calculated_rows:
            for calc_row in variance_model.calculated_rows:
                account_name = calc_row.get('account_name', '')
                values = calc_row.get('values', {})
                period_data = values.get(period, {})

                # Write subtotal name (bolded)
                cell = ws.cell(row=current_row, column=1, value=account_name)
                self.format_bold(ws, f'A{current_row}')

                # Write subtotal values
                budget_value = period_data.get('budget_value')
                actual_value = period_data.get('actual_value')
                dollar_variance = period_data.get('dollar_variance')
                pct_variance = period_data.get('pct_variance')
                is_favorable = period_data.get('is_favorable')
                is_flagged = period_data.get('is_flagged')

                if budget_value is not None:
                    ws.cell(row=current_row, column=2, value=budget_value)
                if actual_value is not None:
                    ws.cell(row=current_row, column=3, value=actual_value)
                if dollar_variance is not None:
                    ws.cell(row=current_row, column=4, value=dollar_variance)
                if pct_variance is not None:
                    ws.cell(row=current_row, column=5, value=pct_variance)

                    # Apply conditional highlighting for unfavorable variances
                    if is_flagged and not is_favorable:
                        abs_pct = abs(pct_variance)
                        if abs_pct > 0.10:  # >10% unfavorable
                            self.apply_conditional_highlight(ws, f'D{current_row}', 'FFC7CE')
                            self.apply_conditional_highlight(ws, f'E{current_row}', 'FFC7CE')
                        elif abs_pct >= 0.05:  # 5-10% unfavorable
                            self.apply_conditional_highlight(ws, f'D{current_row}', 'FFE699')
                            self.apply_conditional_highlight(ws, f'E{current_row}', 'FFE699')

                current_row += 1

        # Apply currency formatting to Budget, Actual, and Variance ($) columns
        if current_row > 2:
            self.format_currency(ws, f'B2:B{current_row - 1}')
            self.format_currency(ws, f'C2:C{current_row - 1}')
            self.format_currency(ws, f'D2:D{current_row - 1}')

        # Apply percentage formatting to Variance (%) column
        if current_row > 2:
            self.format_percentage(ws, f'E2:E{current_row - 1}')

        # Apply borders to entire table
        if current_row > 1:
            self.apply_borders(ws, f'A1:E{current_row - 1}')

        # Auto-adjust column widths
        self.auto_adjust_column_widths(ws)

    def _extract_periods(self, hierarchy: dict) -> list:
        """
        Extract periods from variance hierarchy.

        Args:
            hierarchy: Variance hierarchy tree

        Returns:
            List of period strings found in hierarchy
        """
        periods = set()

        def traverse(node):
            if isinstance(node, dict):
                # Check for 'values' key containing period data
                if 'values' in node:
                    values = node['values']
                    if isinstance(values, dict):
                        periods.update(values.keys())

                # Recurse into children
                if 'children' in node:
                    for child in node['children']:
                        traverse(child)

        # Traverse all top-level sections
        for section in hierarchy.values():
            traverse(section)

        return sorted(list(periods))
