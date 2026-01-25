"""
P&L Forecast Excel writer with three-row confidence intervals.

Generates P&L Forecast sheet with variable-horizon projections (6 or 12 months)
showing Lower Bound, Projected, and Upper Bound as three consecutive rows per metric.
Supports multi-scenario side-by-side display with P&L-specific summary metrics.
"""
from .base_writer import BaseExcelWriter
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


class PLForecastReportWriter(BaseExcelWriter):
    """
    Excel writer for P&L Forecast report with three-row confidence intervals.

    Creates a sheet with variable-horizon monthly columns (6 or 12 months based on
    metadata.forecast_horizon). Each metric displays three consecutive rows:
    Lower Bound (10th percentile), Projected (median), Upper Bound (90th percentile).
    Projected row is bolded. Supports multi-scenario side-by-side comparison.
    Includes P&L-specific summary metrics (gross profit, operating income, net income, margins).
    """

    def write(self, forecast_model) -> None:
        """
        Generate P&L Forecast sheet from PLForecastModel or MultiScenarioForecastResult.

        Args:
            forecast_model: PLForecastModel instance OR MultiScenarioForecastResult with scenarios
        """
        # Create P&L Forecast sheet
        ws = self.workbook.create_sheet('P&L Forecast')

        # Determine if this is multi-scenario or single scenario
        # Strengthen detection to handle Mock objects properly
        is_multi_scenario = (
            hasattr(forecast_model, 'scenarios') and
            isinstance(getattr(forecast_model, 'scenarios', None), list) and
            len(getattr(forecast_model, 'scenarios', [])) > 0
        )

        if is_multi_scenario:
            self._write_multi_scenario(ws, forecast_model)
        else:
            self._write_single_scenario(ws, forecast_model)

        # Auto-adjust column widths
        self.auto_adjust_column_widths(ws)

    def _write_single_scenario(self, ws, forecast_model) -> None:
        """
        Write single scenario P&L Forecast.

        Args:
            ws: Worksheet to write to
            forecast_model: PLForecastModel instance
        """
        # Get forecast horizon from metadata
        forecast_horizon = forecast_model.metadata.get('forecast_horizon', 6)

        # Write header row: Account | Month 1 | Month 2 | ... | Month N
        ws.cell(row=1, column=1, value='Account')
        for month in range(1, forecast_horizon + 1):
            ws.cell(row=1, column=month + 1, value=f'Month {month}')

        # Apply header style
        last_col_letter = get_column_letter(forecast_horizon + 1)
        self.apply_header_style(ws, f'A1:{last_col_letter}1')

        # Track current row
        current_row = 2

        # Traverse hierarchy and write three rows per metric
        # Hierarchy is dict with section keys, iterate over sections
        for section_name, section_node in forecast_model.hierarchy.items():
            for indent_level, name, projected, lower_bound, upper_bound in self.traverse_hierarchy(
                section_node, indent_level=0, forecast=True
            ):
                # Write Lower Bound row
                cell = ws.cell(row=current_row, column=1, value=f'{name} (Lower)')
                cell.alignment = Alignment(indent=indent_level)
                for month in range(1, forecast_horizon + 1):
                    value = lower_bound.get(month)
                    if value is not None:
                        ws.cell(row=current_row, column=month + 1, value=value)
                current_row += 1

                # Write Projected row (bolded)
                cell = ws.cell(row=current_row, column=1, value=f'{name} (Projected)')
                cell.alignment = Alignment(indent=indent_level)
                self.format_bold(ws, f'A{current_row}')
                for month in range(1, forecast_horizon + 1):
                    value = projected.get(month)
                    if value is not None:
                        ws.cell(row=current_row, column=month + 1, value=value)
                # Bold the entire projected row
                self.format_bold(ws, f'B{current_row}:{last_col_letter}{current_row}')
                current_row += 1

                # Write Upper Bound row
                cell = ws.cell(row=current_row, column=1, value=f'{name} (Upper)')
                cell.alignment = Alignment(indent=indent_level)
                for month in range(1, forecast_horizon + 1):
                    value = upper_bound.get(month)
                    if value is not None:
                        ws.cell(row=current_row, column=month + 1, value=value)
                current_row += 1

        # Write calculated rows (Gross Profit, Operating Income, Net Income, margins)
        if forecast_model.calculated_rows:
            # Add blank row for separation
            current_row += 1

            # Define P&L-specific calculated rows in order
            calc_metrics = [
                ('gross_profit', 'Gross Profit', False),
                ('gross_margin_pct', 'Gross Margin %', True),
                ('operating_income', 'Operating Income', False),
                ('operating_margin_pct', 'Operating Margin %', True),
                ('net_income', 'Net Income', False),
            ]

            for calc_key, calc_name, is_percentage in calc_metrics:
                if calc_key in forecast_model.calculated_rows:
                    calc_row_data = forecast_model.calculated_rows[calc_key]

                    projected = calc_row_data.get('projected', {})
                    lower_bound = calc_row_data.get('lower_bound', {})
                    upper_bound = calc_row_data.get('upper_bound', {})

                    # Lower Bound
                    ws.cell(row=current_row, column=1, value=f'{calc_name} (Lower)')
                    for month in range(1, forecast_horizon + 1):
                        value = lower_bound.get(month)
                        if value is not None:
                            ws.cell(row=current_row, column=month + 1, value=value)

                    # Apply percentage formatting if this is a margin metric
                    if is_percentage:
                        self.format_percentage(ws, f'B{current_row}:{last_col_letter}{current_row}')

                    current_row += 1

                    # Projected (bolded)
                    ws.cell(row=current_row, column=1, value=f'{calc_name} (Projected)')
                    self.format_bold(ws, f'A{current_row}')
                    for month in range(1, forecast_horizon + 1):
                        value = projected.get(month)
                        if value is not None:
                            ws.cell(row=current_row, column=month + 1, value=value)
                    self.format_bold(ws, f'B{current_row}:{last_col_letter}{current_row}')

                    # Apply percentage formatting if this is a margin metric
                    if is_percentage:
                        self.format_percentage(ws, f'B{current_row}:{last_col_letter}{current_row}')

                    current_row += 1

                    # Upper Bound
                    ws.cell(row=current_row, column=1, value=f'{calc_name} (Upper)')
                    for month in range(1, forecast_horizon + 1):
                        value = upper_bound.get(month)
                        if value is not None:
                            ws.cell(row=current_row, column=month + 1, value=value)

                    # Apply percentage formatting if this is a margin metric
                    if is_percentage:
                        self.format_percentage(ws, f'B{current_row}:{last_col_letter}{current_row}')

                    current_row += 1

        # Apply currency formatting to all value columns (except percentage rows which were handled separately)
        if current_row > 2:
            self.format_currency(ws, f'B2:{last_col_letter}{current_row - 1}')

        # Apply borders to entire table
        if current_row > 1:
            self.apply_borders(ws, f'A1:{last_col_letter}{current_row - 1}')

    def _write_multi_scenario(self, ws, multi_scenario_result) -> None:
        """
        Write multi-scenario P&L Forecast with scenarios side-by-side.

        Args:
            ws: Worksheet to write to
            multi_scenario_result: MultiScenarioForecastResult instance with scenarios list
        """
        scenarios = multi_scenario_result.scenarios

        # Get forecast horizon from first scenario (all scenarios use same horizon per uniform policy)
        forecast_horizon = scenarios[0].metadata.get('forecast_horizon', 6)

        # Write header row: Account | Scenario1 - Month 1 | Scenario1 - Month 2 | ... | Scenario2 - Month 1 | ...
        ws.cell(row=1, column=1, value='Account')
        col_idx = 2

        for scenario in scenarios:
            scenario_name = scenario.metadata.get('scenario_name', 'Scenario')
            for month in range(1, forecast_horizon + 1):
                ws.cell(row=1, column=col_idx, value=f'{scenario_name} - Month {month}')
                col_idx += 1

        # Apply header style
        last_col_letter = get_column_letter(col_idx - 1)
        self.apply_header_style(ws, f'A1:{last_col_letter}1')

        # Track current row
        current_row = 2

        # Get hierarchy from first scenario to determine structure
        first_scenario = scenarios[0]

        # Traverse hierarchy and write three rows per metric
        # Hierarchy is dict with section keys, iterate over sections
        for section_name, section_node in first_scenario.hierarchy.items():
            for indent_level, name, _, _, _ in self.traverse_hierarchy(
                section_node, indent_level=0, forecast=True
            ):
                # Write Lower Bound row
                cell = ws.cell(row=current_row, column=1, value=f'{name} (Lower)')
                cell.alignment = Alignment(indent=indent_level)
                col_idx = 2
                for scenario in scenarios:
                    # Find matching metric in this scenario's hierarchy
                    lower_bound = self._find_metric_values(scenario.hierarchy, name, 'lower_bound')
                    for month in range(1, forecast_horizon + 1):
                        value = lower_bound.get(month) if lower_bound else None
                        if value is not None:
                            ws.cell(row=current_row, column=col_idx, value=value)
                        col_idx += 1
                current_row += 1

                # Write Projected row (bolded)
                cell = ws.cell(row=current_row, column=1, value=f'{name} (Projected)')
                cell.alignment = Alignment(indent=indent_level)
                self.format_bold(ws, f'A{current_row}')
                col_idx = 2
                for scenario in scenarios:
                    # Find matching metric in this scenario's hierarchy
                    projected = self._find_metric_values(scenario.hierarchy, name, 'projected')
                    for month in range(1, forecast_horizon + 1):
                        value = projected.get(month) if projected else None
                        if value is not None:
                            ws.cell(row=current_row, column=col_idx, value=value)
                        col_idx += 1
                self.format_bold(ws, f'B{current_row}:{last_col_letter}{current_row}')
                current_row += 1

                # Write Upper Bound row
                cell = ws.cell(row=current_row, column=1, value=f'{name} (Upper)')
                cell.alignment = Alignment(indent=indent_level)
                col_idx = 2
                for scenario in scenarios:
                    # Find matching metric in this scenario's hierarchy
                    upper_bound = self._find_metric_values(scenario.hierarchy, name, 'upper_bound')
                    for month in range(1, forecast_horizon + 1):
                        value = upper_bound.get(month) if upper_bound else None
                        if value is not None:
                            ws.cell(row=current_row, column=col_idx, value=value)
                        col_idx += 1
                current_row += 1

        # Write calculated rows if available
        if first_scenario.calculated_rows:
            # Add blank row for separation
            current_row += 1

            # Define P&L-specific calculated rows
            calc_metrics = [
                ('gross_profit', 'Gross Profit', False),
                ('gross_margin_pct', 'Gross Margin %', True),
                ('operating_income', 'Operating Income', False),
                ('operating_margin_pct', 'Operating Margin %', True),
                ('net_income', 'Net Income', False),
            ]

            for calc_key, calc_name, is_percentage in calc_metrics:
                if calc_key in first_scenario.calculated_rows:
                    # Lower Bound
                    ws.cell(row=current_row, column=1, value=f'{calc_name} (Lower)')
                    col_idx = 2
                    for scenario in scenarios:
                        calc_row_data = scenario.calculated_rows.get(calc_key, {})
                        lower_bound = calc_row_data.get('lower_bound', {})
                        for month in range(1, forecast_horizon + 1):
                            value = lower_bound.get(month)
                            if value is not None:
                                ws.cell(row=current_row, column=col_idx, value=value)
                            col_idx += 1
                    current_row += 1

                    # Projected (bolded)
                    ws.cell(row=current_row, column=1, value=f'{calc_name} (Projected)')
                    self.format_bold(ws, f'A{current_row}')
                    col_idx = 2
                    for scenario in scenarios:
                        calc_row_data = scenario.calculated_rows.get(calc_key, {})
                        projected = calc_row_data.get('projected', {})
                        for month in range(1, forecast_horizon + 1):
                            value = projected.get(month)
                            if value is not None:
                                ws.cell(row=current_row, column=col_idx, value=value)
                            col_idx += 1
                    self.format_bold(ws, f'B{current_row}:{last_col_letter}{current_row}')

                    # Apply percentage formatting if this is a margin metric
                    if is_percentage:
                        self.format_percentage(ws, f'B{current_row}:{last_col_letter}{current_row}')

                    current_row += 1

                    # Upper Bound
                    ws.cell(row=current_row, column=1, value=f'{calc_name} (Upper)')
                    col_idx = 2
                    for scenario in scenarios:
                        calc_row_data = scenario.calculated_rows.get(calc_key, {})
                        upper_bound = calc_row_data.get('upper_bound', {})
                        for month in range(1, forecast_horizon + 1):
                            value = upper_bound.get(month)
                            if value is not None:
                                ws.cell(row=current_row, column=col_idx, value=value)
                            col_idx += 1

                    # Apply percentage formatting if this is a margin metric
                    if is_percentage:
                        self.format_percentage(ws, f'B{current_row}:{last_col_letter}{current_row}')

                    current_row += 1

        # Apply currency formatting to all value columns (percentage formatting already applied)
        if current_row > 2:
            self.format_currency(ws, f'B2:{last_col_letter}{current_row - 1}')

        # Apply borders to entire table
        if current_row > 1:
            self.apply_borders(ws, f'A1:{last_col_letter}{current_row - 1}')

    def _find_metric_values(self, hierarchy: dict, metric_name: str, value_type: str) -> dict:
        """
        Find a specific metric's values in hierarchy by name.

        Args:
            hierarchy: Forecast hierarchy tree
            metric_name: Name of the metric to find
            value_type: Type of values to extract ('projected', 'lower_bound', 'upper_bound')

        Returns:
            Dict mapping month numbers to values, or empty dict if not found
        """
        def traverse(node):
            if isinstance(node, dict):
                if node.get('name') == metric_name:
                    return node.get(value_type, {})

                # Recurse into children
                if 'children' in node:
                    for child in node['children']:
                        result = traverse(child)
                        if result:
                            return result

            return {}

        # Search all top-level sections
        for section in hierarchy.values():
            result = traverse(section)
            if result:
                return result

        return {}
